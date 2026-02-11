#!/usr/bin/env python3
"""
Generate Mindforge Student Experience — Project Tracking Spreadsheet.
Run from project root: python docs/planning/generate_tracking_spreadsheet.py
Output: docs/planning/Mindforge_Student_Experience_Project_Tracking.xlsx
"""
import os
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

# All tasks: (Sprint#, SprintName, TaskID, Title, UserStoryShort, Area, Type, AI, Risk, SP, Dependencies)
TASKS = [
    (1, "Foundation & Auth", "1.1", "API gateway and request validation",
     "Backend: REST gateway, validation, consistent error shape", "Backend — API", "Hardening", "Non-AI", "Low", 2, "None"),
    (1, "Foundation & Auth", "1.2", "MPIN verify and token/session issuance",
     "Student: enter 6-digit MPIN, get token/session", "Backend — Auth", "Feature", "Non-AI", "Medium", 3, "1.1"),
    (1, "Foundation & Auth", "1.3", "Lockout status and Forgot MPIN entry points",
     "Student: see lockout status and Forgot MPIN path", "Backend — Auth", "Feature", "Non-AI", "Low", 2, "1.2"),
    (1, "Foundation & Auth", "1.4", "Auth middleware and Bearer token validation",
     "Backend: validate Bearer on protected routes", "Backend — API", "Hardening", "Non-AI", "Low", 1, "1.2"),
    (2, "Data & Backend Core", "2.1", "Database schema and migrations",
     "System: versioned schema for students, activities, etc.", "Backend — Data", "Hardening", "Non-AI", "Low", 3, "None"),
    (2, "Data & Backend Core", "2.2", "Data access layer (repositories/DAOs)",
     "Business layer: repositories for all entities", "Backend — Data", "Feature", "Non-AI", "Low", 3, "2.1"),
    (2, "Data & Backend Core", "2.3", "Business layer skeleton and dependency injection",
     "API layer: business logic calls data access only", "Backend — Business", "Hardening", "Non-AI", "Low", 2, "2.2"),
    (3, "Today's Plan & Activities API", "3.1", "GET /student/today (today's plan)",
     "Student: get today's tasks and progress summary", "Backend — Business", "Feature", "Non-AI", "Low", 2, "2.2, 2.3"),
    (3, "Today's Plan & Activities API", "3.2", "GET /student/activities/:type/:id and POST pause",
     "Student: open activity, see questions, pause/save progress", "Backend — Business", "Feature", "Non-AI", "Low", 2, "2.2, 2.3"),
    (3, "Today's Plan & Activities API", "3.3", "POST /student/activities/:type/:id/respond",
     "Student: submit answer, get correct/incorrect (deterministic grading)", "Backend — Business", "Feature", "Non-AI", "Low", 3, "3.2"),
    (3, "Today's Plan & Activities API", "3.4", "GET /student/results/:type/:id",
     "Student: get activity result (score, breakdown, suggested next)", "Backend — Business", "Feature", "Non-AI", "Low", 1, "3.3"),
    (4, "AI Integration & Grading", "4.1", "AI provider integration and timeout/fallback",
     "Business layer: call AI with timeout and fallback", "Backend — Integration", "Feature", "AI", "Medium", 3, "2.3"),
    (4, "AI Integration & Grading", "4.2", "GET /student/activities/:type/:id/feedback (progressive guidance)",
     "Student: next AI guidance level (Hint → Approach → Concept → Solution)", "Backend — Business", "Feature", "AI", "Medium", 3, "3.3, 4.1"),
    (4, "AI Integration & Grading", "4.3", "Deterministic and AI-assisted grading rules",
     "System: deterministic for objective; AI/rubric for open-ended", "Backend — Business", "Feature", "AI", "Low", 2, "3.3, 4.1"),
    (5, "Attendance & Doubts API", "5.1", "GET /student/attendance (summary and calendar)",
     "Student: attendance summary and calendar for period", "Backend — Business", "Feature", "Non-AI", "Low", 2, "2.2, 2.3"),
    (5, "Attendance & Doubts API", "5.2", "Doubts: GET list, GET thread, POST message with syllabus context",
     "Student: doubt threads with Class→Subject→Chapter→Topic context", "Backend — Business", "Feature", "AI", "Medium", 3, "4.1, 2.2, 2.3"),
    (5, "Attendance & Doubts API", "5.3", "GET /student/profile and GET /student/sync/status",
     "Student: profile and sync status (conflict hint)", "Backend — Business", "Feature", "Non-AI", "Low", 2, "2.2, 2.3"),
    (6, "Client — Login, Home, Activity, Results", "6.1", "Client app shell and API client (auth, REST, error shape)",
     "Client: Bearer token, error shape, retry", "Frontend", "Feature", "Non-AI", "Low", 2, "API contract"),
    (6, "Client — Login, Home, Activity, Results", "6.2", "Login screen (Screen 1) per UX spec",
     "Student: 6-digit MPIN, keypad, error/lockout states", "Frontend", "Feature", "Non-AI", "Low", 2, "6.1; API 1.2, 1.3"),
    (6, "Client — Login, Home, Activity, Results", "6.3", "Home screen (Screen 2) — Today's Plan",
     "Student: task cards, progress bar, completed today", "Frontend", "Feature", "Non-AI", "Low", 2, "6.1; API 3.1"),
    (6, "Client — Login, Home, Activity, Results", "6.4", "Activity screen (Screen 3) — question, answer, AI feedback panel",
     "Student: question, submit answer, AI feedback or fallback", "Frontend", "Feature", "AI", "Medium", 3, "6.1; API 3.2, 3.3, 4.2"),
    (6, "Client — Login, Home, Activity, Results", "6.5", "Results screen (Screen 4)",
     "Student: score, breakdown, suggested next", "Frontend", "Feature", "Non-AI", "Low", 1, "6.1; API 3.4"),
    (7, "Client — Attendance, Doubts, Profile & Sync", "7.1", "Attendance screen (Screen 5) — calendar and summary",
     "Student: attendance summary and calendar", "Frontend", "Feature", "Non-AI", "Low", 2, "6.1; API 5.1"),
    (7, "Client — Attendance, Doubts, Profile & Sync", "7.2", "Doubts screen (Screen 6) with syllabus context selector",
     "Student: AI chat with Class→Subject→Chapter→Topic context", "Frontend", "Feature", "AI", "Medium", 3, "6.1; API 5.2"),
    (7, "Client — Attendance, Doubts, Profile & Sync", "7.3", "Profile screen and logout",
     "Student: profile (name, class, board, progress) and logout", "Frontend", "Feature", "Non-AI", "Low", 1, "6.1; API 5.3"),
    (7, "Client — Attendance, Doubts, Profile & Sync", "7.4", "Sync status and conflict modal",
     "Student: sync status; conflict 'Use latest?' / 'Keep current'", "Frontend", "Feature", "Non-AI", "Low", 2, "6.1, 6.3; API 5.3"),
    (8, "DevOps, Security Hardening & NotebookLM", "8.1", "Deployment pipeline and single-region (India) hosting",
     "Team: CI/CD, deploy staging→production, single region", "DevOps", "Hardening", "Non-AI", "Low", 2, "App, DB"),
    (8, "DevOps, Security Hardening & NotebookLM", "8.2", "Secrets vault and no secrets in code/client",
     "System: API keys, DB credentials in vault", "DevOps / Security", "Hardening", "Non-AI", "Medium", 1, "8.1"),
    (8, "DevOps, Security Hardening & NotebookLM", "8.3", "Rate limiting and audit logging",
     "System: rate limit on auth; audit log for sensitive ops", "Backend — Security", "Hardening", "Non-AI", "Low", 2, "1.2, 2.1"),
    (8, "DevOps, Security Hardening & NotebookLM", "8.4", "Monitoring and alerting",
     "Team: health checks, latency/error metrics, alert on 5xx/auth", "DevOps", "Hardening", "Non-AI", "Low", 1, "8.1"),
    (8, "DevOps, Security Hardening & NotebookLM", "8.5", "NotebookLM read-only integration (teaching feed)",
     "Business layer: consume daily teaching feed for homework/quiz", "Backend — Integration", "Feature", "Non-AI", "Medium", 2, "2.2, 2.3"),
]

STATUS_OPTIONS = "Not Started,In Progress,Code Complete,Review,Done"

def main():
    script_dir = Path(__file__).resolve().parent
    out_path = script_dir / "Mindforge_Student_Experience_Project_Tracking.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Sprint & Task Tracking"

    # Header row
    headers = [
        "Sprint #", "Sprint Name", "Task ID", "Task Title", "User Story (short)",
        "Area", "Type", "AI/Non-AI", "Risk", "SP", "Status", "Start Date", "Completed Date",
        "Dependencies", "Acceptance Criteria Ref", "Tracker Notes (Dev Agent)"
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="748B75", end_color="748B75", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data rows
    for row_idx, t in enumerate(TASKS, 2):
        sprint_num, sprint_name, task_id, title, user_story, area, typ, ai, risk, sp, deps = t
        ws.cell(row=row_idx, column=1, value=sprint_num)
        ws.cell(row=row_idx, column=2, value=sprint_name)
        ws.cell(row=row_idx, column=3, value=task_id)
        ws.cell(row=row_idx, column=4, value=title)
        ws.cell(row=row_idx, column=5, value=user_story)
        ws.cell(row=row_idx, column=6, value=area)
        ws.cell(row=row_idx, column=7, value=typ)
        ws.cell(row=row_idx, column=8, value=ai)
        ws.cell(row=row_idx, column=9, value=risk)
        ws.cell(row=row_idx, column=10, value=sp)
        ws.cell(row=row_idx, column=11, value="Not Started")  # Status
        ws.cell(row=row_idx, column=12, value="")  # Start Date
        ws.cell(row=row_idx, column=13, value="")  # Completed Date
        ws.cell(row=row_idx, column=14, value=deps)
        ws.cell(row=row_idx, column=15, value=f"Backlog § Task {task_id} Checklist")
        ws.cell(row=row_idx, column=16, value="")  # Tracker Notes

    # Data validation for Status (column K = 11)
    dv = DataValidation(
        type="list",
        formula1=f'"{STATUS_OPTIONS}"',
        allow_blank=False
    )
    dv.error = "Pick one of: Not Started, In Progress, Code Complete, Review, Done"
    ws.add_data_validation(dv)
    dv.add(f"K2:K{len(TASKS) + 1}")  # Status = column K (11)

    # Column widths
    widths = [8, 22, 8, 42, 38, 18, 10, 8, 8, 4, 14, 12, 14, 18, 28, 32]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = min(w, 50)

    # Freeze header and first columns
    ws.freeze_panes = "F2"

    # Sprint summary sheet
    ws2 = wb.create_sheet("Sprint Summary", 1)
    ws2.cell(row=1, column=1, value="Sprint #")
    ws2.cell(row=1, column=2, value="Sprint Name")
    ws2.cell(row=1, column=3, value="Capacity (SP)")
    ws2.cell(row=1, column=4, value="Planned SP")
    ws2.cell(row=1, column=5, value="Tasks")
    ws2.cell(row=1, column=6, value="Done")
    ws2.cell(row=1, column=7, value="Remaining")
    for c in range(1, 8):
        ws2.cell(row=1, column=c).font = Font(bold=True)
        ws2.cell(row=1, column=c).fill = PatternFill(start_color="748B75", end_color="748B75", fill_type="solid")
        ws2.cell(row=1, column=c).font = Font(bold=True, color="FFFFFF")
    sprint_info = [
        (1, "Foundation & Auth", 10, 8, 4),
        (2, "Data & Backend Core", 10, 8, 3),
        (3, "Today's Plan & Activities API", 10, 8, 4),
        (4, "AI Integration & Grading", 10, 8, 3),
        (5, "Attendance & Doubts API", 10, 7, 3),
        (6, "Client — Login, Home, Activity, Results", 10, 10, 5),
        (7, "Client — Attendance, Doubts, Profile & Sync", 10, 8, 4),
        (8, "DevOps, Security Hardening & NotebookLM", 10, 8, 5),
    ]
    for row_idx, (snum, sname, cap, planned, n_tasks) in enumerate(sprint_info, 2):
        ws2.cell(row=row_idx, column=1, value=snum)
        ws2.cell(row=row_idx, column=2, value=sname)
        ws2.cell(row=row_idx, column=3, value=cap)
        ws2.cell(row=row_idx, column=4, value=planned)
        ws2.cell(row=row_idx, column=5, value=n_tasks)
        ws2.cell(row=row_idx, column=6, value=0)  # Done — update as Dev completes
        ws2.cell(row=row_idx, column=7, value=f"=E{row_idx}-F{row_idx}")
    ws2.column_dimensions["A"].width = 8
    ws2.column_dimensions["B"].width = 38
    for c in "CDEFG":
        ws2.column_dimensions[c].width = 12

    wb.save(out_path)
    print(f"Saved: {out_path}")
    return out_path

if __name__ == "__main__":
    main()
