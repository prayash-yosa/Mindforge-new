#!/usr/bin/env python3
"""
Generate Mindforge Admin App — Project Tracking Spreadsheet.
Run: python docs/planning/generate_admin_tracking_spreadsheet.py
Output: docs/planning/Mindforge_Admin_App_Project_Tracking.xlsx
"""
import csv
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

HEADER_COLOR = "748B75"
STATUS_OPTIONS = "Not Started,In Progress,Code Complete,Review,Done"

SPRINT_INFO = [
    (1, "Workspace & Foundation", 10, 8, 4),
    (2, "User Approval & Lifecycle", 10, 8, 3),
    (3, "Fees Config & Payment Info", 10, 8, 3),
    (4, "Payment Entry & Ledger", 10, 8, 3),
    (5, "Frontend Shell, Dashboard & Users", 10, 8, 3),
    (6, "Frontend Fees & Payments", 10, 8, 3),
    (7, "Frontend Payment Info & Audit", 10, 8, 3),
    (8, "Hardening & Validation", 10, 8, 3),
]


def main():
    script_dir = Path(__file__).resolve().parent
    csv_path = script_dir / "Mindforge_Admin_App_Project_Tracking.csv"
    out_path = script_dir / "Mindforge_Admin_App_Project_Tracking.xlsx"

    rows = []
    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(row)

    wb = Workbook()
    ws = wb.active
    ws.title = "Project Tracking"

    headers = [
        "Sprint #", "Sprint Name", "Task ID", "Task Title", "User Story (short)",
        "Area", "Type", "AI/Non-AI", "Risk", "SP", "Status", "Start Date",
        "Completed Date", "Dependencies", "Acceptance Criteria Ref",
        "Tracker Notes (Dev Agent)",
    ]

    header_fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    csv_keys = [
        "Sprint #", "Sprint Name", "Task ID", "Task Title", "User Story (short)",
        "Area", "Type", "AI/Non-AI", "Risk", "SP", "Status", "Start Date",
        "Completed Date", "Dependencies", "Acceptance Criteria Ref",
        "Tracker Notes (Dev Agent)",
    ]

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, key in enumerate(csv_keys, 1):
            val = row_data.get(key, "")
            if key == "Sprint #" or key == "SP":
                try:
                    val = int(val)
                except (ValueError, TypeError):
                    pass
            ws.cell(row=row_idx, column=col_idx, value=val)

    dv = DataValidation(type="list", formula1=f'"{STATUS_OPTIONS}"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(f"K2:K{len(rows) + 1}")

    widths = [8, 22, 8, 42, 38, 22, 10, 8, 8, 4, 14, 12, 14, 18, 28, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = min(w, 50)

    ws.freeze_panes = "F2"

    wb.save(out_path)
    print(f"Saved: {out_path}")
    return out_path


if __name__ == "__main__":
    main()
