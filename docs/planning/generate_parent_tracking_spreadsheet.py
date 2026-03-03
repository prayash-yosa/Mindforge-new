#!/usr/bin/env python3
"""
Generate Mindforge Parent App — Project Tracking Spreadsheet with Charts.
Run from project root: python docs/planning/generate_parent_tracking_spreadsheet.py
Output: docs/planning/Mindforge_Parent_App_Project_Tracking.xlsx
"""
import csv
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.chart import PieChart, BarChart, Reference
    from openpyxl.chart.label import DataLabelList
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.chart import PieChart, BarChart, Reference
    from openpyxl.chart.label import DataLabelList


HEADER_COLOR = "748B75"
STATUS_OPTIONS = "Not Started,In Progress,Code Complete,Review,Done"

SPRINT_INFO = [
    (1, "Workspace & Foundation", 10, 8, 4),
    (2, "Auth & Linkage", 10, 8, 3),
    (3, "Read APIs", 10, 8, 3),
    (4, "Dashboard Aggregation", 10, 8, 3),
    (5, "Frontend Shell, Login & Home", 10, 8, 3),
    (6, "Frontend Progress & Attendance", 10, 8, 3),
    (7, "Frontend Fees & Profile", 10, 8, 3),
    (8, "Hardening & Validation", 10, 8, 3),
]


def main():
    script_dir = Path(__file__).resolve().parent
    csv_path = script_dir / "Mindforge_Parent_App_Project_Tracking.csv"
    out_path = script_dir / "Mindforge_Parent_App_Project_Tracking.xlsx"

    rows = []
    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(row)

    wb = Workbook()

    # ── Sheet 1: Sprint & Task Tracking ─────────────────────────
    ws = wb.active
    ws.title = "Project Tracking"

    headers = [
        "Sprint #", "Sprint Name", "Task ID", "Task Title", "User Story (short)",
        "Area", "Type", "AI/Non-AI", "Risk", "SP", "Status", "Start Date",
        "Completed Date", "Dependencies", "Acceptance Criteria Ref",
        "Tracker Notes (Dev Agent)",
    ]

    header_fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    csv_keys = [
        "Sprint #", "Sprint Name", "Task ID", "Task Title", "User Story (short)",
        "Area", "Type", "AI/Non-AI", "Risk", "SP", "Status", "Start Date",
        "Completed Date", "Dependencies", "Acceptance Criteria Ref",
        "Tracker Notes (Dev Agent)",
    ]

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, key in enumerate(csv_keys, 1):
            val = row_data.get(key, "")
            if key == "Sprint #" or key == "SP":
                try:
                    val = int(val)
                except (ValueError, TypeError):
                    pass
            ws.cell(row=row_idx, column=col_idx, value=val)

    dv = DataValidation(type="list", formula1=f'"{STATUS_OPTIONS}"', allow_blank=False)
    dv.error = "Pick one of: Not Started, In Progress, Code Complete, Review, Done"
    ws.add_data_validation(dv)
    dv.add(f"K2:K{len(rows) + 1}")

    widths = [8, 22, 8, 42, 38, 22, 10, 8, 8, 4, 14, 12, 14, 18, 28, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = min(w, 50)

    ws.freeze_panes = "F2"

    # ── Sheet 2: Sprint Summary ──────────────────────────────────
    ws2 = wb.create_sheet("Sprint Summary")

    summary_headers = ["Sprint #", "Sprint Name", "Capacity (SP)", "Planned SP", "Tasks", "Done", "Remaining"]
    for c, h in enumerate(summary_headers, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    done_by_sprint = {}
    for row_data in rows:
        try:
            snum = int(row_data.get("Sprint #", 0))
        except (ValueError, TypeError):
            continue
        if row_data.get("Status", "").strip() == "Done":
            done_by_sprint[snum] = done_by_sprint.get(snum, 0) + 1

    for row_idx, (snum, sname, cap, planned, n_tasks) in enumerate(SPRINT_INFO, 2):
        done = done_by_sprint.get(snum, 0)
        ws2.cell(row=row_idx, column=1, value=snum)
        ws2.cell(row=row_idx, column=2, value=sname)
        ws2.cell(row=row_idx, column=3, value=cap)
        ws2.cell(row=row_idx, column=4, value=planned)
        ws2.cell(row=row_idx, column=5, value=n_tasks)
        ws2.cell(row=row_idx, column=6, value=done)
        ws2.cell(row=row_idx, column=7, value=n_tasks - done)

    ws2.column_dimensions["A"].width = 8
    ws2.column_dimensions["B"].width = 42
    for c in "CDEFG":
        ws2.column_dimensions[c].width = 14

    last_summary_row = len(SPRINT_INFO) + 1

    # ── Charts ───────────────────────────────────────────────────

    total_tasks = len(rows)
    total_done = sum(1 for r in rows if r.get("Status", "").strip() == "Done")
    total_remaining = total_tasks - total_done

    chart_data_row = last_summary_row + 3
    ws2.cell(row=chart_data_row, column=1, value="Category")
    ws2.cell(row=chart_data_row, column=2, value="Count")
    ws2.cell(row=chart_data_row + 1, column=1, value="Done")
    ws2.cell(row=chart_data_row + 1, column=2, value=total_done)
    ws2.cell(row=chart_data_row + 2, column=1, value="Remaining")
    ws2.cell(row=chart_data_row + 2, column=2, value=total_remaining)

    pie = PieChart()
    pie.title = "Overall Task Completion"
    pie.width = 16
    pie.height = 12
    labels = Reference(ws2, min_col=1, min_row=chart_data_row + 1, max_row=chart_data_row + 2)
    data = Reference(ws2, min_col=2, min_row=chart_data_row, max_row=chart_data_row + 2)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showVal = True
    from openpyxl.chart.series import DataPoint
    pt_done = DataPoint(idx=0)
    pt_done.graphicalProperties.solidFill = "4A7C59"
    pt_remain = DataPoint(idx=1)
    pt_remain.graphicalProperties.solidFill = "D4A574"
    pie.series[0].data_points = [pt_done, pt_remain]
    ws2.add_chart(pie, "I2")

    bar = BarChart()
    bar.type = "col"
    bar.title = "Sprint Progress (Done vs Remaining)"
    bar.width = 22
    bar.height = 12
    bar.y_axis.title = "Tasks"
    bar.x_axis.title = "Sprint"
    cats = Reference(ws2, min_col=2, min_row=2, max_row=last_summary_row)
    done_data = Reference(ws2, min_col=6, min_row=1, max_row=last_summary_row)
    remain_data = Reference(ws2, min_col=7, min_row=1, max_row=last_summary_row)
    bar.add_data(done_data, titles_from_data=True)
    bar.add_data(remain_data, titles_from_data=True)
    bar.set_categories(cats)
    bar.series[0].graphicalProperties.solidFill = "4A7C59"
    bar.series[1].graphicalProperties.solidFill = "D4A574"
    ws2.add_chart(bar, "I18")

    sp_bar = BarChart()
    sp_bar.type = "col"
    sp_bar.title = "Story Points by Sprint"
    sp_bar.width = 22
    sp_bar.height = 12
    sp_bar.y_axis.title = "Story Points"
    sp_bar.x_axis.title = "Sprint"
    sp_cats = Reference(ws2, min_col=2, min_row=2, max_row=last_summary_row)
    sp_data = Reference(ws2, min_col=4, min_row=1, max_row=last_summary_row)
    sp_bar.add_data(sp_data, titles_from_data=True)
    sp_bar.set_categories(sp_cats)
    sp_bar.series[0].graphicalProperties.solidFill = "748B75"
    ws2.add_chart(sp_bar, "I34")

    wb.save(out_path)
    print(f"Saved: {out_path}")
    print(f"  Total tasks: {total_tasks}")
    print(f"  Done: {total_done}")
    print(f"  Remaining: {total_remaining}")
    return out_path


if __name__ == "__main__":
    main()
